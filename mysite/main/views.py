from django.db.models.functions import TruncDate
from django.shortcuts import render,redirect
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseRedirect
from django.urls import reverse
from .models import *
from django.contrib.auth.forms import UserCreationForm
import pandas as pd
from datetime import datetime, date
from urllib.parse import urlencode
from django.db.models import Q
from django.core.paginator import Paginator
from django.core.mail import send_mail
from .filters import PedidoFilter
from .forms import *
from django.contrib.auth.models import Group
from django.contrib.auth import get_user
from . import views
from django.contrib.auth import authenticate, login, logout
from django_tables2 import SingleTableMixin
from django_filters.views import FilterView
from django.conf import settings
from .tables import UtilizadoresTable
from .filters import UtilizadoresFilter
from dateutil.parser import parse
from django.db.models import Count, Avg
from django.utils.timezone import datetime, timedelta
import re
<<<<<<< HEAD
import datetime
from datetime import datetime, timedelta
=======
import smtplib
from smtplib import SMTPException
from openpyxl import Workbook
from openpyxl.styles import Font

>>>>>>> 3f70598d08a8bd65e43e51907fa2fc999a749706

def user_check(request, user_profile = None):
    ''' 
    Verifica se o utilizador que esta logado pertence a pelo menos um dos perfis mencionados 
    e.g. user_profile = {Administrador,Coordenador,ProfessorUniversitario}
    Isto faz com que o user que esta logado possa ser qualquer um dos 3 perfis. 
    '''
    if not request.user.is_authenticated:
        return {'exists': False, 'render': redirect('main:login')}
    elif user_profile is not None:
        matches_profile = False
        for profile in user_profile:
            if profile.objects.filter(utilizador_ptr_id = request.user.id).exists():
                return {'exists': True, 'firstProfile': profile.objects.filter(utilizador_ptr_id = request.user.id).first()}
        return {'exists': False, 
                'render': render(request=request,
                            template_name='mensagem.html',
                            context={
                                'tipo':'error',
                                'm':'Não tem permissões para aceder a esta página!'
                            })
                }
    raise Exception('Unknown Error!')

### Pagina Inicial ###
def homepage(request):
   return render(request, template_name="main/inicio.html")

### App ###
def homepage5(request):
   return render(request, template_name="main/app.html")

### Registo ###
def register(request):
   return render(request, template_name="users/criar_utilizador.html")

### Criação de Pedidos ###
def PedidoHorarios(request):
   name=Docente.objects.all()
   UC = UnidadesCurriculares.objects.all()
   user = get_user(request)
   if request.method == "POST":
      uc_list = request.POST.getlist('unc')
      dia = request.POST['data']
      assunto = request.POST['assunto']
      desc = request.POST['desc']
      descri_list = request.POST.getlist('descri')
      hora_inicio_list = request.POST.getlist('hora_inicio')
      hora_fim_list = request.POST.getlist('hora_fim')
      tarefa = request.POST.getlist('tarefa')
      dia2 =request.POST.getlist('data2')

      # Verifica se a data escolhida é anterior ao dia de hoje
      if parse(dia).date() < date.today():
         error = 'A data escolhida é anterior ao dia de hoje!'
         return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name,"UC": UC})
      
      if Pedido.objects.filter(assunto=assunto, dia=dia).exists():
            # Se já existir, retorne uma mensagem de erro
            error = 'Já existe um pedido com o mesmo assunto!'
            return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name,"UC": UC})
      for hora_inicio, hora_fim in zip(hora_inicio_list, hora_fim_list):
         try:
            parse(hora_inicio).time()
            parse(hora_fim).time()
         except ValueError:
            error = 'Formato de hora inválido!'
            return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name,"UC": UC})
      # Verifica se há um ano letivo ativo
      try:
            ano_letivo_ativo = AnoLetivo.objects.get(ativo=True)
      except AnoLetivo.DoesNotExist:
            error = 'É necessário ter um ano letivo ativo.'
            return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name, "UC": UC})

        
      docente = Docente.objects.get(utilizador_ptr=user)
      new_pedido = Pedido(assunto=assunto,desc=desc,dia=dia, tipo="Horário",Docente=docente,Anoletivo=ano_letivo_ativo)
      new_pedido.save()
      
      for i in range(len(uc_list)):
         new_pedido_hor = PedidoHorario(
            uc=uc_list[i],
            hora_inicio=hora_inicio_list[i],
            hora_fim=hora_fim_list[i],
            descri=descri_list[i],
            pedido=new_pedido,
            tarefa=tarefa[i],
            dia=dia2[i]
         )
         if PedidoHorario.objects.filter(uc=uc_list[i], dia=dia2[i],hora_inicio=hora_inicio_list[i],hora_fim=hora_fim_list[i],tarefa=tarefa[i]).exists():
            error = 'Pedido de Horário igual!'
            ultimo_pedido = Pedido.objects.last()
            ultimo_pedido.delete()
            return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name,"UC": UC})
         if parse(dia2[i]).date() < date.today():
            error = 'A data escolhida é anterior ao dia de hoje!'
            ultimo_pedido = Pedido.objects.last()
            ultimo_pedido.delete()
            return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name,"UC": UC})
         hora_inicio_list[i] = parse(hora_inicio_list[i]).time()
         hora_fim_list[i] = parse(hora_fim_list[i]).time()
         if hora_inicio_list[i] >= hora_fim_list[i]:
            error = 'A hora de fim deve ser maior que a hora de início!'
            ultimo_pedido = Pedido.objects.last()
            ultimo_pedido.delete()
            return render(request, 'main/PedidoHorario.html', {"error": error, "nome": name,"UC": UC})
         new_pedido_hor.save()
      

      pending_pedidos = Pedido.objects.filter(Docente=docente).exclude(status__in=['Concluido']).count()
      # Calculate the average number of pedidos processed per day in the last 7 days
      seven_days_ago = datetime.now().date() - timedelta(days=7)
      avg_processed_per_day = Pedido.objects.filter(Docente=docente, dia__gte=seven_days_ago).annotate(
         date=TruncDate('dia')
      ).values('date').annotate(processed_count=Count('id')).aggregate(avg_processed=Avg('processed_count'))


      succes = 'Enviado para a base de dados'
      return render(request, template_name="main/PedidoHorario.html",context={"nome": name,"UC": UC,"succes": succes, "pending_pedidos": pending_pedidos,
         "avg_processed_per_day": avg_processed_per_day['avg_processed']})
  
   return render(request, template_name="main/PedidoHorario.html",context={"nome": name,"UC": UC})


def PedidosOUT(request):
    user = get_user(request)
    name = Docente.objects.all()
    if request.method == "POST":
        assunto = request.POST['assunto']
        descricao = request.POST['desc']
        data = request.POST['data']
        arquivo = request.FILES.getlist('resume')

        # Verifica se a data escolhida é anterior ao dia de hoje
        if parse(data).date() < date.today():
          error = 'A data escolhida é anterior ao dia de hoje!'
          return render(request, 'main/PedidosOutros.html', {"error": error, "nome": name})
        
        if Pedido.objects.filter(assunto=assunto, dia=data).exists():
            # Se já existir, retorne uma mensagem de erro
            error = 'Já existe um pedido com o mesmo assunto!'
            return render(request, 'main/PedidosOutros.html', {"error": error, "nome": name})

        docente = Docente.objects.get(utilizador_ptr=user)
        new_Pedido = Pedido(assunto=assunto, desc=descricao, dia=data, tipo = "Outros",Docente=docente)
        new_Pedido.save()

        for i in range(len(arquivo)):
           new_pedido_outros = PedidosOutros(
            arquivo = arquivo[i],
            pedido = new_Pedido   
           )
           new_pedido_outros.save()
        pending_pedidos = Pedido.objects.filter(Docente=docente).exclude(status__in=['Concluido']).count()
      # Calculate the average number of pedidos processed per day in the last 7 days
        seven_days_ago = datetime.now().date() - timedelta(days=7)
        avg_processed_per_day = Pedido.objects.filter(Docente=docente, dia__gte=seven_days_ago).annotate(
         date=TruncDate('dia')
      ).values('date').annotate(processed_count=Count('id')).aggregate(avg_processed=Avg('processed_count'))


        succes = 'Enviado para a base de dados'
        return render(request, template_name="main/PedidosOutros.html",context={"nome": name,"succes": succes, "pending_pedidos": pending_pedidos,
         "avg_processed_per_day": avg_processed_per_day['avg_processed']})


    return render(request, template_name="main/PedidosOutros.html", context={"nome": name})


def PedidoUnidadeCurricular(request):
      user = get_user(request)
      UC = UnidadesCurriculares.objects.all()
      if request.method == "POST":
         date = request.POST['data']
         assunto = request.POST['assunto']
         desc = request.POST['desc']

         uc_list = request.POST.getlist('unc')
         tarefa = request.POST.getlist('tarefa')
         regente = request.POST.getlist('regente')
         descri = request.POST.getlist('descri')
         docente = Docente.objects.get(utilizador_ptr=user)
         new_Pedido = Pedido(assunto=assunto,desc=desc,dia=date,tipo="Unidade Curricular",Docente=docente)
         new_Pedido.save()

         
         for i in range(len(uc_list)):
            new_PedidoUC = PedidoUC(
               uc=uc_list[i],
               tarefa=tarefa[i],
               regente=regente[i],
               descri=descri[i], 
               pedido = new_Pedido
            )
            new_PedidoUC.save()
      return render(request, template_name="main/PedidoUC.html",context={"UC": UC})


def PedidoSalas(request):
   user = get_user(request)
   Salas = Sala.objects.all()
   Edificios = Edificio.objects.all()
   UC = UnidadesCurriculares.objects.all()
   if request.method == "POST":
   
      salaa= request.POST.getlist('salaa')
      uc_list = request.POST.getlist('unc')
      dia= request.POST['data']
      assunto = request.POST['assunto']

      hora_de_inicio_list = request.POST.getlist('hora_inicio')
      hora_de_fim_list= request.POST.getlist('hora_fim')
      desc = request.POST['desc']
      descri_list = request.POST.getlist('descri')
      tarefa = request.POST.getlist('tarefa')
      dia2 =request.POST.getlist('data2')

      # Verifica se a data escolhida é anterior ao dia de hoje
      if parse(dia).date() < date.today():
         error = 'A data escolhida é anterior ao dia de hoje!'
         return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas , "edificios": Edificios,"UC": UC})
      
      if Pedido.objects.filter(assunto=assunto, dia=dia).exists():
            # Se já existir, retorne uma mensagem de erro
            error = 'Já existe um pedido com o mesmo assunto!'
            return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas , "edificios": Edificios,"UC": UC})
      for hora_de_inicio, hora_de_fim in zip(hora_de_inicio_list, hora_de_fim_list):
         try:
            parse(hora_de_inicio).time()
            parse(hora_de_fim).time()
         except ValueError:
            error = 'Formato de hora inválido!'
            return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas , "edificios": Edificios,"UC": UC})
            # Verifica se há um ano letivo ativo
      try:
            ano_letivo_ativo = AnoLetivo.objects.get(ativo=True)
      except AnoLetivo.DoesNotExist:
            error = 'É necessário ter um ano letivo ativo.'
            return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas, "edificios": Edificios, "UC": UC})

      docente = Docente.objects.get(utilizador_ptr=user)
      new_Pedido = Pedido(assunto = assunto, desc = desc, dia = dia, tipo = "Sala",Docente=docente, Anoletivo=ano_letivo_ativo)
      new_Pedido.save()

      for i in range(len(uc_list)):
         new_PedidoSala = PedidoSala(
            
            hora_de_inicio=hora_de_inicio_list[i],
            hora_de_fim=hora_de_fim_list[i],
            descri=descri_list[i],
            uc = uc_list[i],
            dia = dia2[i],
            pedido=new_Pedido
         )  

         if PedidoSala.objects.filter(uc=uc_list[i], dia=dia2[i],hora_de_inicio=hora_de_inicio_list[i],hora_de_fim=hora_de_fim_list[i],tarefa=tarefa[i]).exists():
            error = 'Pedido de Horário igual!'
            ultimo_pedido = Pedido.objects.last()
            ultimo_pedido.delete()
            return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas , "edificios": Edificios,"UC": UC})
         if parse(dia2[i]).date() < date.today():
            error = 'A data escolhida é anterior ao dia de hoje!'
            ultimo_pedido = Pedido.objects.last()
            ultimo_pedido.delete()
            return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas , "edificios": Edificios,"UC": UC})
         hora_de_inicio_list[i] = parse(hora_de_inicio_list[i]).time()
         hora_de_fim_list[i] = parse(hora_de_fim_list[i]).time()
         if hora_de_inicio_list[i] >= hora_de_fim_list[i]:
            error = 'A hora de fim deve ser maior que a hora de início!'
            ultimo_pedido = Pedido.objects.last()
            ultimo_pedido.delete()
            return render(request, 'main/PedidoSala.html', {"error": error, "salaa": Salas , "edificios": Edificios,"UC": UC})
         new_PedidoSala.save()
      
      succes = 'Enviado para a base de dados'
      return render(request, template_name="main/PedidoSala.html",context={ "salaa": Salas , "edificios": Edificios,"UC": UC, "succes": succes})
  
   return render(request, template_name="main/PedidoSala.html",context={ "salaa": Salas , "edificios": Edificios,"UC": UC})

### Update dos Pedidos ###
def updateHorario(request, pk):
    pedido = Pedido.objects.get(id=pk)
    pedido_horario = PedidoHorario.objects.filter(pedido=pk).all()
    UC = UnidadesCurriculares.objects.all()
    if request.method == "POST":
        pedido.dia = request.POST['data']
        dias = pedido.dia
        pedido.assunto = request.POST['assunto']
        pedido.desc = request.POST['desc']

        if datetime.strptime(dias, '%Y-%m-%d').date() < date.today():
         error = 'A data escolhida é anterior ao dia de hoje!'
         return render(request, 'main/PedidoHorario2.html', {"error": error,"assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_horario":pedido_horario,
        "UC": UC,})
      
        if Pedido.objects.filter(Q(assunto=pedido.assunto, dia=pedido.dia) & ~Q(id=pk)).exists():
            # Se já existir, retorne uma mensagem de erro
            error = 'Já existe um pedido com o mesmo assunto!'
            return render(request, 'main/PedidoHorario2.html', {"error": error,"assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_horario":pedido_horario,
        "UC": UC,})
      
        pedido.save()

        uc_list = request.POST.getlist('unc')
        descri_list = request.POST.getlist('descri')
        hora_inicio_list = request.POST.getlist('hora_inicio')
        hora_fim_list = request.POST.getlist('hora_fim')
        tarefa = request.POST.getlist('tarefa')
        dia2 =request.POST.getlist('data2')

        
        # Combine as listas em uma lista de tuplas
        updates = zip(pedido_horario, uc_list, descri_list, hora_inicio_list, hora_fim_list, tarefa, dia2)

        # Atualize os objetos PedidoHorario
        for pedido_horario, uc, descri, hora_inicio, hora_fim, tarefa, dia2 in updates:
            pedido_horario.uc = uc
            pedido_horario.descri = descri
            pedido_horario.hora_inicio = hora_inicio
            pedido_horario.hora_fim = hora_fim
            pedido_horario.tarefa = tarefa
            pedido_horario.dia = dia2
            
            try:
                  datetime.strptime(pedido_horario.hora_inicio, '%H:%M:%S')
                  datetime.strptime(pedido_horario.hora_fim, '%H:%M:%S')
            except ValueError:
                  error = 'Formato de hora inválido!'
                  return render(request, 'main/PedidoHorario2.html', {"error": error,"assunto": pedido.assunto,
            "desc": pedido.desc,
            "dia": pedido.dia,
            "pedido_horario":pedido_horario,
            "UC": UC,})
            pedido_horario.save()
        
        redirect_url = reverse('main:tablePedidos')
        params = urlencode({'success': 'Enviado para a base de dados'})
        redirect_url = f"{redirect_url}?{params}"
        return HttpResponseRedirect(redirect_url)
    return render(request, template_name="main/PedidoHorario2.html", context={
        "assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_horario":pedido_horario,
        "UC": UC,
    })


def UPDATEPeidosOUT(request, pk):
    pedido = Pedido.objects.get(id=pk)
    pedido_outros = PedidosOutros.objects.filter(pedido=pk).all()
    if request.method == "POST":
        pedido.dia = request.POST['data']
        dias = pedido.dia
        pedido.assunto = request.POST['assunto']
        pedido.desc = request.POST['desc']
        if datetime.strptime(dias, '%Y-%m-%d').date() < date.today():
            error = 'A data escolhida é anterior ao dia de hoje!'
            return render(request, 'main/PedidosOutros2.html', {"error": error, "assunto": pedido.assunto,
                                                                "desc": pedido.desc,
                                                                "dia": pedido.dia,
                                                                "pedido_outros": pedido_outros,})
        if Pedido.objects.filter(Q(assunto=pedido.assunto, dia=pedido.dia) & ~Q(id=pk)).exists():
            error = 'Já existe um pedido com o mesmo assunto!'
            return render(request, 'main/PedidosOutros2.html', {"error": error, "assunto": pedido.assunto,
                                                                "desc": pedido.desc,
                                                                "dia": pedido.dia,
                                                                "pedido_outros": pedido_outros,})

        pedido.save()
        arquivo_list = request.POST.getlist("arquivo")

        # Combine the pedido_outros objects and arquivo_list into a list of tuples
        updates = zip(pedido_outros, arquivo_list)

        # Update the PedidosOutros objects
        for pedido_outros, arquivo in updates:
            pedido_outros.arquivo = arquivo
            pedido_outros.save()

        redirect_url = reverse('main:tablePedidos')
        params = urlencode({'success': 'Enviado para a base de dados'})
        redirect_url = f"{redirect_url}?{params}"
        return HttpResponseRedirect(redirect_url)

    # Construct a list of arquivo names for the pedido_outros objects
    arquivo_names = [pedido_outro.arquivo.name if pedido_outro.arquivo else '' for pedido_outro in pedido_outros]

    return render(request, template_name="main/PedidosOutros2.html", context={
        "assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_outros": pedido_outros,
        "arquivo_names": arquivo_names,
    })


def updateUC(request, pk):
    pedido = Pedido.objects.get(id=pk)
    pedido_horario = PedidoUC.objects.filter(pedido=pk).all()
    UC = UnidadesCurriculares.objects.all()
    if request.method == "POST":
        pedido.dia = request.POST['data']
        dias = pedido.dia
        pedido.assunto = request.POST['assunto']
        pedido.desc = request.POST['desc']

        if datetime.strptime(dias, '%Y-%m-%d').date() < date.today():
         error = 'A data escolhida é anterior ao dia de hoje!'
         return render(request, 'main/PedidoUC2.html', {"error": error,"assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_horario":pedido_horario,
        "UC": UC,})
      
        if Pedido.objects.filter(Q(assunto=pedido.assunto, dia=pedido.dia) & ~Q(id=pk)).exists():
            # Se já existir, retorne uma mensagem de erro
            error = 'Já existe um pedido com o mesmo assunto!'
            return render(request, 'main/PedidoUC2.html', {"error": error,"assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_horario":pedido_horario,
        "UC": UC,})
      
        pedido.save()

        uc_list = request.POST.getlist('unc')
        descri_list = request.POST.getlist('descri')
        tarefa = request.POST.getlist('tarefa')
        regente = request.POST.getlist('regente')

        
        # Combine as listas em uma lista de tuplas
        updates = zip(pedido_horario, uc_list, regente,descri_list, tarefa)

        # Atualize os objetos PedidoHorario
        for pedido_horario, uc, descri_list, regente, tarefa in updates:
            pedido_horario.uc = uc
            pedido_horario.descri = descri_list
            pedido_horario.tarefa = tarefa
            pedido_horario.regente = regente
            pedido_horario.save()
            
        
        redirect_url = reverse('main:tablePedidos')
        params = urlencode({'success': 'Enviado para a base de dados'})
        redirect_url = f"{redirect_url}?{params}"
        return HttpResponseRedirect(redirect_url)
    return render(request, template_name="main/PedidoUC2.html", context={
        "assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "pedido_horario":pedido_horario,
        "UC": UC,
    })



def updateSala(request, pk):
    pedido = Pedido.objects.get(id=pk)
    pedido_Sala = PedidoSala.objects.filter(pedido=pk).all()
    UC = UnidadesCurriculares.objects.all()
    Salas = Sala.objects.all()
    if request.method == "POST":
      pedido.dia = request.POST['data']
      pedido.assunto = request.POST['assunto']
      pedido.desc = request.POST['desc']

      if datetime.strptime(pedido.dia, '%Y-%m-%d').date() < date.today():
         error = 'A data escolhida é anterior ao dia de hoje!'
         return render(request, 'main/PedidoSala2.html', {"error": error,"assunto": pedido.assunto,
         "desc": pedido.desc,
         "dia": pedido.dia,
         "Salas":Salas,
         "pedido_Sala":pedido_Sala, "UC":UC})
      
      if Pedido.objects.filter(Q(assunto=pedido.assunto, dia=pedido.dia) & ~Q(id=pk)).exists():
         # Se já existir, retorne uma mensagem de erro
         error = 'Já existe um pedido com o mesmo assunto!'
         return render(request, 'main/PedidoSala2.html', {"error": error,"assunto": pedido_Sala.assunto,
         "desc": pedido_Sala.desc,
         "dia": pedido_Sala.dia,
         "Salas":Salas,
         "pedido_Sala":pedido_Sala, "UC":UC})
                                                      
      pedido.save()

      sala_list = request.POST.getlist('sala')
      descricao_list = request.POST.getlist('descri')
      hora_inicio_list = request.POST.getlist('hora_de_inicio')
      hora_fim_list = request.POST.getlist('hora_de_fim')
      tarefa = request.POST.getlist('tarefa')
      dia2 =request.POST.getlist('data2')

      # Combine as listas em uma lista de tuplas
      updates = zip(pedido_Sala, sala_list, descricao_list, hora_inicio_list, hora_fim_list, tarefa, dia2)

      # Atualize os objetos PedidoSala
      for pedido_Sala, sala, descri, hora_inicio, hora_fim, tarefa, dia2 in updates:
         pedido_Sala.sal = sala
         pedido_Sala.descri = descri
         pedido_Sala.hora_de_inicio = hora_inicio
         pedido_Sala.hora_de_fim = hora_fim
         pedido_Sala.tarefa = tarefa
         pedido_Sala.dia = dia2
         try:
            datetime.strptime(pedido_Sala.hora_de_inicio, '%H:%M:%S')
            datetime.strptime(pedido_Sala.hora_de_fim, '%H:%M:%S')
         except ValueError:
            error = 'Formato de hora inválido!'
            return render(request, 'main/PedidoSala2.html', {"error": error,"assunto": pedido.assunto,
         "desc": pedido.desc,
         "dia": pedido.dia,
         "Salas":Salas,
         "pedido_Sala":pedido_Sala, "UC":UC})
         pedido_Sala.save()
      
      redirect_url = reverse('main:tablePedidos')
      params = urlencode({'success': 'Enviado para a base de dados'})
      redirect_url = f"{redirect_url}?{params}"
      return HttpResponseRedirect(redirect_url)
    return render(request, template_name="main/PedidoSala2.html", context={
        "assunto": pedido.assunto,
        "desc": pedido.desc,
        "dia": pedido.dia,
        "Salas":Salas,
        "pedido_Sala":pedido_Sala, "UC":UC     
    })

### Apagar Pedidos ISTO SO JA APAGA TODOS OS TIPOS DO PEDIDO ###
def deletHorario(request,pk):
   PedidosHor = Pedido.objects.get(id=pk)
   if request.method == "POST":
      PedidosHor.delete()
      return redirect('/tablePedidos')
   return render(request, template_name="main/deleteH.html",context={'item': PedidosHor})



### Tabela Pedidos###
def tablePedidos(request):
   if request.user.is_authenticated:
        user = get_user(request)
        if user.groups.filter(name="Administrador").exists():
            u = "Administrador"
        elif user.groups.filter(name="Docente").exists():
            u = "Docente"
        elif user.groups.filter(name="Funcionário").exists():
            u = "Funcionario"
        else:
            u = ""
   else:
        u = ""
   pedidoshorario = Pedido.objects.all()
   pedidoshorarios = PedidoHorario.objects.all()
   funciona = Funcionario.objects.all()
   myFilter = PedidoFilter(request.GET,queryset=pedidoshorario)
   pedidoshorario = myFilter.qs
   if user.groups.filter(name="Docente").exists():
      docente = Docente.objects.get(utilizador_ptr=user)
      pedidoshorario = pedidoshorario.filter(Docente=docente)
   paginator = Paginator(pedidoshorario, 10)  # Define a quantidade de itens por página
   page_number = request.GET.get('page')  # Obtém o número da página da solicitação
   page_obj = paginator.get_page(page_number)
   success = request.GET.get('success')
   

   if request.method == 'POST':
      pedido_id = request.POST.get('pedido_id') 
      pedido = Pedido.objects.get(id=pedido_id)
      
      if request.POST.get('desassociar'):
         pedido.Funcionario = None
         pedido.atribuido = "Não Atribuido"
         pedido.status = "Registado"
         pedido.save()
         return redirect(f'{reverse("main:tablePedidos")}?page={page_obj.number}&success=Funcionario desassociado!')
      else:
         user = get_user(request)
         funcionario = Funcionario.objects.get(utilizador_ptr=user)
         pedido.status = "Em Análise"
         pedido.atribuido = "Atribuido"
         pedido.Funcionario = funcionario
         pedido.save()
         return redirect(f'{reverse("main:tablePedidos")}?page={page_obj.number}&success=Funcionario Atribuido!')
      
      
  
   return render(request, template_name="main/tableHorario.html",context={"Pedido":page_obj, "item":pedidoshorarios, "funcio":funciona,'success': success, 'myFilter':myFilter, 'u':u})


### Tabela Pedidos veie###
def tablePedidos2(request,pk):
   try:
    pedido = Pedido.objects.get(id=pk)
    if PedidoHorario.objects.filter(pedido=pedido).exists():
        pedidoshorario = PedidoHorario.objects.filter(pedido=pedido)
        return render(request, template_name="main/tableHorario2.html", context={"Pedido": pedidoshorario,"pedido":pedido})
    elif PedidosOutros.objects.filter(pedido=pedido).exists():
        pedidosoutros = PedidosOutros.objects.filter(pedido=pedido)
        return render(request, template_name="main/tableHorario2.html", context={"Pedido": pedidosoutros,"pedido":pedido})
    elif PedidoSala.objects.filter(pedido=pedido).exists():
        pedidossala = PedidoSala.objects.filter(pedido=pedido)
        return render(request, template_name="main/tableHorario2.html", context={"Pedido": pedidossala,"pedido":pedido})
    elif PedidoUC.objects.filter(pedido=pedido).exists():
        pedidosuc = PedidoUC.objects.filter(pedido=pedido)
        return render(request, template_name="main/tableHorario2.html", context={"Pedido": pedidosuc, "pedido":pedido})
    
   except Pedido.DoesNotExist:
      return HttpResponseNotFound('Pedido não encontrado')
      
   pedidoshorario = PedidoHorario.objects.filter(pedido=pedido)
   return render(request, template_name="main/tableHorario2.html",context={"Pedido":pedidoshorario,"pedido":pedido})




def AnoLetivoAdd(request):
   UC = UnidadesCurriculares.objects.all()
   success = None
   error = None

   # Gerar a lista de opções de anos letivos
   anos_letivos = [f"{year}/{year+1:02}" for year in range(date.today().year, 2099)]

   if request.method == "POST":
      anoletivo = request.POST['anoletivo']
      datainicio = request.POST['datainicio']
      datafinal = request.POST['datafinal']

      # Verificar se o ano letivo já existe
      if AnoLetivo.objects.filter(anoletivo=anoletivo).exists():
         error = f"O ano letivo {anoletivo} já existe."
         return render(request, template_name="main/PedidoAL.html",context={"UC": UC,"anos_letivos": anos_letivos,'succes':success,'error':error})
      else:
         # Verificar se as datas pertencem ao ano letivo escolhido
         ano_inicio = int(anoletivo.split('/')[0])
         ano_fim = int(anoletivo.split('/')[1])
         if int(datainicio.split('-')[0]) != ano_inicio or int(datafinal.split('-')[0]) != ano_fim:
            error = "As datas devem pertencer ao ano letivo selecionado."
            return render(request, template_name="main/PedidoAL.html",context={"UC": UC,"anos_letivos": anos_letivos,'succes':success,'error':error})
         elif datainicio >= datafinal:
            error = "A data de início deve ser anterior à data de fim."
            return render(request, template_name="main/PedidoAL.html",context={"UC": UC,"anos_letivos": anos_letivos,'succes':success,'error':error})
         else:
            new_ano = AnoLetivo(
               anoletivo=anoletivo,
               datainicio=datainicio,
               datafinal=datafinal
            )
            new_ano.save()
            success = f"Ano letivo {anoletivo} criado com sucesso."

   return render(request, template_name="main/PedidoAL.html", context={"UC": UC, "anos_letivos": anos_letivos, "succes": success, "error": error})

### Update Ano Letivo ###
def updateAL(request, pk):
    pedido = AnoLetivo.objects.get(id=pk)
    UC = UnidadesCurriculares.objects.all()
    success = None
    error = None

   # Gerar a lista de opções de anos letivos
    anos_letivos = [f"{year}/{year+1:02}" for year in range(date.today().year, 2099)]
    if request.method == "POST":
        pedido.anoletivo = request.POST['anoletivo']
        pedido.datainicio = request.POST['datainicio']
        pedido.datafinal = request.POST['datafinal']
        if AnoLetivo.objects.exclude(id=pk).filter(anoletivo=pedido.anoletivo).exists():
          error = f"O ano letivo {pedido.anoletivo} já existe."
          return render(request, template_name="main/PedidoAL2.html",context={"anoletivo": pedido.anoletivo,"datainicio": pedido.datainicio,"datafinal": pedido.datafinal,"anos_letivos": anos_letivos,'succes':success,'error':error})
        else:
         # Verificar se as datas pertencem ao ano letivo escolhido
         ano_inicio = int(pedido.anoletivo.split('/')[0])
         ano_fim = int(pedido.anoletivo.split('/')[1])
         if int(pedido.datainicio.split('-')[0]) != ano_inicio or int(pedido.datafinal.split('-')[0]) != ano_fim:
            error = "As datas devem pertencer ao ano letivo selecionado."
            return render(request, template_name="main/PedidoAL2.html",context={"anoletivo": pedido.anoletivo,"datainicio": pedido.datainicio,"datafinal": pedido.datafinal,"anos_letivos": anos_letivos,'succes':success,'error':error})
         elif pedido.datainicio >= pedido.datafinal:
            error = "A data de início deve ser anterior à data de fim."
            return render(request, template_name="main/PedidoAL2.html",context={"anoletivo": pedido.anoletivo,"datainicio": pedido.datainicio,"datafinal": pedido.datafinal,"anos_letivos": anos_letivos,'succes':success,'error':error})
         else:
          pedido.save()
          return redirect('main:tableAL')
    return render(request, template_name="main/PedidoAL2.html", context={"anoletivo": pedido.anoletivo,"datainicio": pedido.datainicio,"datafinal": pedido.datafinal,"anos_letivos": anos_letivos,'succes':success,'error':error
      
    })


### Apagar Ano Letivo ###
def deletAL(request,pk):
   PedidosAL = AnoLetivo.objects.get(id=pk)
   if request.method == "POST":
      PedidosAL.delete()
      return redirect('/tableAL')
   return render(request, template_name="main/deleteAL.html",context={'item': PedidosAL})


### Tabela Ano Letivo ###
def tableAL(request):
    pedidosAL = AnoLetivo.objects.all()

    # Verificar se há um ano letivo ativo
    ano_letivo_ativo = AnoLetivo.objects.filter(ativo=True).first()



    if request.method == "POST":
        ano_id = request.POST.get("ano_id")

        if ano_id:
            # Verificar se o valor de ano_id é um número válido
            try:
                ano_id = int(ano_id)
                ano_letivo = AnoLetivo.objects.get(id=ano_id)
                ano_letivo.ativo = True
                ano_letivo.save()
            except (ValueError, AnoLetivo.DoesNotExist):
                # Lidar com o valor inválido de ano_id ou se não existir um ano letivo com o ID fornecido
                pass

    return render(request, template_name="main/tableAL.html", context={"AnoLetivo": pedidosAL})


### Importações ###
def uploadRUC(request):
    if request.method == 'POST':
        if not request.FILES:
            error = 'Selecione um arquivo para fazer o upload'
            return render(request=request, template_name='main/IMPORTCSV.html', context={'error': error})
        if not request.FILES['file'].name.endswith('.xlsx'):
            error = 'O arquivo deve estar no formato .xlsx'
            return render(request=request, template_name='main/IMPORTCSV.html', context={'error': error})
        excel_file = request.FILES['file']
        excel_data = pd.ExcelFile(excel_file)
        sheetx = pd.read_excel(excel_data, sheet_name=0)
        testes = sheetx[['Ano letivo', 'Docente', 'Regência', 'Tipo', 'Horas']] 

        for index, row in testes.iterrows():
            AnoLetivo = row['Ano letivo']
            nome_docente = row['Docente'].split("-")[1].strip()
            Docente = nome_docente
            Regencia = row['Regência']
            Tipo = row['Tipo']
            Horas = row['Horas']
            #Codigo = row['Codigo']
            #Desc = row['Descrição']

            
            # Verificar se a linha começa com "ENGENHARIA"
            if not Regencia.startswith('ENGENHARIA') and not re.search(r'FCT', Regencia):
                # Extrair apenas os nomes das unidades curriculares
                match = re.search(r'(.+?)\s*\(\d+\)', Regencia)
                if match:
                    Regencia = match.group(1).strip()
                
                match_tipo = re.search(r'Regência de disciplina', Tipo)
                if match_tipo:
                    Tipo = match_tipo.group(0)
                
                    novAUC = UnidadesCurriculares(AnoLetivo=AnoLetivo, Docentes=Docente, Regência=Regencia, Tipo=Tipo, Horas=Horas)
                    novAUC.save()
            succes = 'Importado para a base de dados'
        return render(request=request, template_name='main/IMPORTCSV.html', context={"succes": succes})
    return render(request=request, template_name='main/IMPORTCSV.html')


def uploadDSD(request):
    if request.method == 'POST':
        if not request.FILES:
            error = 'Selecione um arquivo para fazer o upload'
            return render(request=request, template_name='main/upload_DSD.html', context={'error': error})
        if not request.FILES['file'].name.endswith('.xls'):
            error = 'O arquivo deve estar no formato .xls'
            return render(request=request, template_name='main/upload_DSD.html', context={'error': error})
        excel_file = request.FILES['file']
        excel_data = pd.ExcelFile(excel_file)
        sheetx = pd.read_excel(excel_data, sheet_name=0)
        testes = sheetx[['Período', 'Cód. disciplina', 'Disciplina', 'Inst. disciplina', 'Turma', 'Curso',
                         'Cód. Docente', 'Docente', 'Depart. docente', 'Horas semanais' , 'Data início', 'Data fim']] 

        for index, row in testes.iterrows():
            Periodo = row['Período']
            codDisci = row['Cód. disciplina']
            disciplina = row['Disciplina']
            instituto = row['Inst. disciplina']
            turma = row['Turma']
            curso = row['Curso']
            codDocente = row['Cód. Docente']
            docente = row['Docente']
            departDocente = row['Depart. docente']
            horasSem = row['Horas semanais']
            Datainicial = row['Data início']
            DataFim = row['Data fim']
            novAUC = DSD(Periodo = Periodo, codDisci =  codDisci, disciplina = disciplina, instituto = instituto,
                         turma = turma, curso = curso, codDocente = codDocente, docente = docente, departDocente = departDocente,
                         horasSem = horasSem, Datainicial = Datainicial, DataFim = DataFim)
            novAUC.save()
        success = 'Dados do DSD importados com sucesso'
        return render(request=request, template_name='main/upload_DSD.html', context={'success': success})
    return render(request=request, template_name='main/upload_DSD.html')



def uploadDocente(request):
    if request.method == 'POST':
        if not request.FILES:
            error = 'Selecione um arquivo para fazer o upload'
            return render(request=request, template_name='main/Upload_Docentes.html', context={'error': error})
        if not request.FILES['file'].name.endswith('.xls'):
            error = 'O arquivo deve estar no formato .xls'
            return render(request=request, template_name='main/Upload_Docentes.html', context={'error': error})
        excel_file = request.FILES['file']
        excel_data = pd.ExcelFile(excel_file)
        sheetx = pd.read_excel(excel_data, sheet_name=0)
        testes = sheetx[['Código','Ativo', 'Nome', 'Indivíduo', 'Data de nascimento', 'Sexo', 'Tipo de identificação', 'Identificação', 'Data de emissão da identificação', 'Nacionalidade', 'Arquivo', 'Data de validade da identificação', 'NIF', 'País fiscal', 'Digito verificação', '&nbsp;']] 

        for index, row in testes.iterrows():
            codigo = row['Código']
            ativo = row['Ativo']
            nome = row['Nome']
            if ativo == "S":
                    ativo = True
            else: ativo = False

            novoDocente = Docente_import(codigo=codigo, ativo=ativo, nome=nome)
            novoDocente.save()
            succes = 'Importado para a base de dados'
        return render(request=request, template_name='main/Upload_Docentes.html', context={"succes": succes})
    return render(request=request, template_name='main/Upload_Docentes.html')

def uploadSALAS(request):
    if request.method == 'POST':
        if not request.FILES:
            error = 'Selecione um arquivo para fazer o upload'
            return render(request=request, template_name='main/upload_SALAS.html', context={'error': error})
        if not request.FILES['file'].name.endswith('.xlsx'):
            error = 'O arquivo deve estar no formato .xlsx'
            return render(request=request, template_name='main/upload_SALAS.html', context={'error': error})
        excel_file = request.FILES['file']
        excel_data = pd.ExcelFile(excel_file)
        sheetx = pd.read_excel(excel_data, sheet_name=0)
        testes = sheetx[['Nome Instituição', 'Desc. Edifício', 'Desc. Sala', 'Des. Categoria', 'Id. tipo sala', 'Lotação presencial sala']] 

        for index, row in testes.iterrows():
            NomeInstituição = row['Nome Instituição']
            DescEdifício = row['Desc. Edifício']
            DescSala = row['Desc. Sala']
            DesCategoria = row['Des. Categoria']
            IdTipoSala = row['Id. tipo sala']
            LotaçãoPresencialSala = row['Lotação presencial sala']

            novAUC = Sala(NomeInstituição=NomeInstituição, DescEdifício=DescEdifício, DescSala=DescSala, DesCategoria=DesCategoria, IdTipoSala=IdTipoSala, LotaçãoPresencialSala=LotaçãoPresencialSala)
            novAUC.save()
        return redirect('main:Upload_Salas')
    return render(request=request, template_name='main/Upload_Salas.html')


### Estatisticas ###
def obter_dados_pedidos_funcionarios():
    funcionarios = Funcionario.objects.all()

    pedidos_recebidos = []

    for funcionario in funcionarios:
        num_pedidos_recebidos = Pedido.objects.filter(Funcionario=funcionario).count()
        pedidos_recebidos.append(num_pedidos_recebidos)

    return funcionarios, pedidos_recebidos

###quantidade Pedidos ####
def tableEstatisticaPedido(request):
   
       # Verificar se os registros já existem
   if not EstatisticaPedido.objects.filter(Status='Em Análise').exists():
        EstatisticaPedido.objects.create(Status='Em Análise')
   if not EstatisticaPedido.objects.filter(Status='Registado').exists():
        EstatisticaPedido.objects.create(Status='Registado')
   if not EstatisticaPedido.objects.filter(Status='Concluído').exists():
        EstatisticaPedido.objects.create(Status='Concluído')
   
       # Get all the employees
   funcionarios = Funcionario.objects.all()

   tempos_processamento = []
   for funcionario in funcionarios:
        pedidos_funcionario = Pedido.objects.filter(Funcionario=funcionario, status='Concluído')
        total_tempos = 0
        num_pedidos_concluidos = pedidos_funcionario.count()
        for pedido in pedidos_funcionario:
            # Calculate the processing time for each order
            dias_processamento = (pedido.diaFinal - pedido.diaCriado).days + 1
            total_tempos += dias_processamento
        if num_pedidos_concluidos > 0:
            media_tempos = total_tempos / num_pedidos_concluidos
        else:
            media_tempos = 0
        tempos_processamento.append(media_tempos)

   pedidos_concluidos = []
   for funcionario in funcionarios:
        # Count the number of completed orders for each employee
        num_pedidos_concluidos = Pedido.objects.filter(Funcionario=funcionario, status='Concluído').count()
        pedidos_concluidos.append(num_pedidos_concluidos)
   
   num_pedidos = Pedido.objects.all().count()
   # Obter o número de pedidos com o status "Em Análise"
   num_pedidos_em_analise = Pedido.objects.filter(status='Em Análise').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Em Análise')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()

   # Obter o número de pedidos com o status "Concluido"
   num_pedidos_em_analise = Pedido.objects.filter(status='Concluído').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Concluido')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()

   # Obter o número de pedidos com o status "Rejeitado"
   num_pedidos_em_analise = Pedido.objects.filter(status='Registado').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Registado')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()



   pedidosSala = EstatisticaPedido.objects.all()
   return render(request, template_name="main/quantidade_pedido.html",context={"Pedido":pedidosSala, "Funcionarios": funcionarios, "PedidosConcluidos": pedidos_concluidos, "TemposProcessamento": tempos_processamento})



###Quantidade de pedidos processados pelo funcionario ####
def estatistica_pedido_processado_funcionario(request):
   
       # Verificar se os registros já existem
   if not EstatisticaPedido.objects.filter(Status='Em Análise').exists():
        EstatisticaPedido.objects.create(Status='Em Análise')
   if not EstatisticaPedido.objects.filter(Status='Registado').exists():
        EstatisticaPedido.objects.create(Status='Registado')
   if not EstatisticaPedido.objects.filter(Status='Concluído').exists():
        EstatisticaPedido.objects.create(Status='Concluído')
   

    # Definir a data de início e fim dos últimos 30 dias
   data_fim = datetime.now().date()
   data_inicio = data_fim - timedelta(days=29)

    # Obter todos os funcionários
   funcionarios = Funcionario.objects.all()

   tempos_processamento = []
   pedidos_concluidos = []

   for funcionario in funcionarios:
        # Filtrar pedidos concluídos do funcionário nos últimos 30 dias
        pedidos_funcionario = Pedido.objects.filter(Funcionario=funcionario, status='Concluído', diaCriado__range=[data_inicio, data_fim])
        
        total_tempos = 0
        num_pedidos_concluidos = pedidos_funcionario.count()

        for pedido in pedidos_funcionario:
            # Calcular o tempo de processamento para cada pedido
            dias_processamento = (pedido.diaFinal - pedido.diaCriado).days + 1
            total_tempos += dias_processamento
        
        if num_pedidos_concluidos > 0:
            media_tempos = total_tempos / num_pedidos_concluidos
        else:
            media_tempos = 0

        tempos_processamento.append(media_tempos)
        pedidos_concluidos.append(num_pedidos_concluidos)
   
   num_pedidos = Pedido.objects.all().count()
   # Obter o número de pedidos com o status "Em Análise"
   num_pedidos_em_analise = Pedido.objects.filter(status='Em Análise').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Em Análise')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()

   # Obter o número de pedidos com o status "Concluido"
   num_pedidos_em_analise = Pedido.objects.filter(status='Concluído').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Concluido')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()

   # Obter o número de pedidos com o status "Rejeitado"
   num_pedidos_em_analise = Pedido.objects.filter(status='Registado').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Registado')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()



   pedidosSala = EstatisticaPedido.objects.all()
   return render(request, template_name="main/pedidos_processados_funcionario.html",context={"Pedido":pedidosSala, "Funcionarios": funcionarios, "PedidosConcluidos": pedidos_concluidos, "TemposProcessamento": tempos_processamento})



### Tempo de Processamento ###

###quantidade Pedidos ####
def estatistica_tempo_processado(request):
   
       # Verificar se os registros já existem
   if not EstatisticaPedido.objects.filter(Status='Em Análise').exists():
        EstatisticaPedido.objects.create(Status='Em Análise')
   if not EstatisticaPedido.objects.filter(Status='Registado').exists():
        EstatisticaPedido.objects.create(Status='Registado')
   if not EstatisticaPedido.objects.filter(Status='Concluído').exists():
        EstatisticaPedido.objects.create(Status='Concluído')
   
    # Definir a data de início e fim dos últimos 30 dias
   data_fim = datetime.now().date()
   data_inicio = data_fim - timedelta(days=29)

    # Obter todos os funcionários
   funcionarios = Funcionario.objects.all()

   tempos_processamento = []
   pedidos_concluidos = []

   for funcionario in funcionarios:
        # Filtrar pedidos concluídos do funcionário nos últimos 30 dias
        pedidos_funcionario = Pedido.objects.filter(Funcionario=funcionario, status='Concluído', diaCriado__range=[data_inicio, data_fim])
        
        total_tempos = 0
        num_pedidos_concluidos = pedidos_funcionario.count()

        for pedido in pedidos_funcionario:
            # Calcular o tempo de processamento para cada pedido
            dias_processamento = (pedido.diaFinal - pedido.diaCriado).days + 1
            total_tempos += dias_processamento
        
        if num_pedidos_concluidos > 0:
            media_tempos = total_tempos / num_pedidos_concluidos
        else:
            media_tempos = 0

        tempos_processamento.append(media_tempos)
        pedidos_concluidos.append(num_pedidos_concluidos)

   num_pedidos = Pedido.objects.all().count()
   
   num_pedidos = Pedido.objects.all().count()
   # Obter o número de pedidos com o status "Em Análise"
   num_pedidos_em_analise = Pedido.objects.filter(status='Em Análise').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Em Análise')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()

   # Obter o número de pedidos com o status "Concluido"
   num_pedidos_em_analise = Pedido.objects.filter(status='Concluído').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Concluido')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()

   # Obter o número de pedidos com o status "Rejeitado"
   num_pedidos_em_analise = Pedido.objects.filter(status='Registado').count()
   estatistica_pedido_em_analise = EstatisticaPedido.objects.get(Status='Registado')
   estatistica_pedido_em_analise.NmrPedido = num_pedidos_em_analise
   porcentagem_analise = (num_pedidos_em_analise / num_pedidos) * 100
   texto_formatado = '{:.1f}%'.format(porcentagem_analise)
   estatistica_pedido_em_analise.percetagem =texto_formatado
   estatistica_pedido_em_analise.save()



   pedidosSala = EstatisticaPedido.objects.all()
   return render(request, template_name="main/tempo_processado.html",context={"Pedido":pedidosSala, "Funcionarios": funcionarios, "PedidosConcluidos": pedidos_concluidos, "TemposProcessamento": tempos_processamento})


### Coisa na minha opiniao nao são necessarias ###
def meus(request):
   pedidosoutro = PedidosOutros.objects.all()
   return render(request, template_name="main/bulma_table.html",context={"Pedido":pedidosoutro})


def deletOutros(request,pk):
   PedidosOut = PedidosOutros.objects.get(id=pk)
   if request.method == "POST":
      PedidosOut.delete()
      return redirect('/meus')
   return render(request, template_name="main/delete.html",context={'item': PedidosOut})


def tableHorario2(request, pk):
   try:
        pedido = Pedido.objects.get(id=pk)
        pedidoshorario = PedidoHorario.objects.filter(pedido=pedido)
   except Pedido.DoesNotExist:
        return HttpResponseNotFound('Pedido não encontrado')
   return render(request, template_name="main/tableHorario2.html",context={"Pedido":pedidoshorario})





def tableUC(request):
   pedidosUC = PedidoUC.objects.all()
   return render(request, template_name="main/tableUC.html",context={"Pedido":pedidosUC})


def deletUC(request,pk):
   PedidosUC = PedidoUC.objects.get(id=pk)
   if request.method == "POST":
      PedidosUC.delete()
      return redirect('/tableUC')
   return render(request, template_name="main/deleteUC.html",context={'item': PedidosUC})


def tableSala(request):
   pedidosSala = PedidoSala.objects.all()
   return render(request, template_name="main/tableSala.html",context={"Pedido":pedidosSala})


def deletSala(request,pk):
   PedidosSala = PedidoSala.objects.get(id=pk)
   if request.method == "POST":
      PedidosSala.delete()
      return redirect('/tableSala')
   return render(request, template_name="main/deleteS.html",context={'item': PedidosSala})



#Login e Register
def login_action(request):
    ''' Fazer login na plataforma do dia aberto e gestão de acessos à plataforma '''
    if request.user.is_authenticated: 
        return redirect("main:home")   
    else:
        u=""
    msg=False
    error=""
    if request.method == 'POST':
        form = LoginForm(request=request, data=request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        if username=="" or password=="":
                msg=True
                error="Todos os campos são obrigatórios!"
        else:
            user = authenticate(username=username, password=password)
            if user is not None:
                utilizador = Utilizador.objects.get(id=user.id)
                if utilizador.valido=="False": 
                    msg=True
                    error="O seu registo ainda não foi validado"
                elif utilizador.valido=="Rejeitado":
                    msg=True
                    error="O seu registo não é válido"
                else:
                    login(request, user)
                    return redirect('main:mensagem',1)
            else:
                msg=True
                error="O nome de utilizador ou a palavra-passe inválidos!"
    form = LoginForm()
    return render(request=request,
                  template_name="main/login.html",
                  context={"form": form,"msg": msg, "error": error, 'u': u})

def logout_action(request):
    ''' Fazer logout na plataforma '''
    logout(request)
    return redirect('main:home')

def escolher(request):
    ''' Escolher tipo de perfil para criar um utilizador '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Administrador").exists():
            u = "Administrador"
        elif user.groups.filter(name = "Funcionário").exists():
            u = "Funcionário"
        elif user.groups.filter(name = "Docente").exists():
            u = "Docente" 
        else:
            u=""     
    else:
        u=""
    utilizadores = ["Funcionário","Docente", "Administrador"]
    return render(request=request, template_name='main/escolher_perfil.html', context={"utilizadores": utilizadores,'u': u})


def register(request, id):
    ''' Criar um novo utilizador que poderá ter de ser validado dependendo do seu tipo '''
    faculdade = Faculdade.objects.all()
    departamento = Departamento.objects.all()
    if request.user.is_authenticated:
        user = get_user(request)
        if user.groups.filter(name="Administrador").exists():
            u = "Administrador"
        elif user.groups.filter(name="Docente").exists():
            u = "Docente"
        elif user.groups.filter(name="Funcionário").exists():
            u = "Funcionario"
        else:
            u = ""
    else:
        u = ""
    msg = False
    if request.method == "POST":
        tipo = id
        if tipo == 1:
            form = FuncionarioRegisterForm(request.POST)
            perfil = "Funcionario"
            my_group, _ = Group.objects.get_or_create(name='Funcionário')
        elif tipo == 2:
            form = DocenteRegisterForm(request.POST)
            perfil = "Docente"
            my_group, _ = Group.objects.get_or_create(name='Docente')
        elif tipo == 3:
            form = AdministradorRegisterForm(request.POST)
            perfil = "Administrador"
            my_group, _ = Group.objects.get_or_create(name='Administrador')
        else:
            return redirect("main:escolher-perfil")

        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')
            first_name = form.cleaned_data.get('first_name')
            user.groups.add(my_group)

            if tipo == 3:
                user.valido = True
                user.save()
                p = 1
            else:
                user.valido = False
                recipient_id = user.id
                user.save()
                p = 0
               # views.enviar_notificacao_automatica(request, "validarRegistosPendentes", recipient_id)
            if request.user.is_authenticated:
                user = get_user(request)
                if user.groups.filter(name="Administrador").exists():
                    return redirect("main:concluir-registo", 2)
            else:
                return redirect("main:concluir-registo", p)
        else:
            msg = True
            tipo = id
            return render(request=request,
                          template_name="main/criar_utilizador.html",
                          context={"form": form, 'perfil': perfil, 'u': u, 'registo': tipo, 'msg': msg, 'faculdade' : faculdade, 'departamento':departamento})
    else:
        tipo = id
        if tipo == 1:
            form = FuncionarioRegisterForm()
            perfil = "Funcionário"
        elif tipo == 2:
            form = DocenteRegisterForm()
            perfil = "Docente"
        elif tipo == 3:
            form = AdministradorRegisterForm()
            perfil = "Administrador"
        else:
            return redirect("main:escolher-perfil")
    return render(request=request,
                  template_name="main/criar_utilizador.html",
                  context={"form": form, 'perfil': perfil, 'u': u, 'registo': tipo, 'msg': msg, 'faculdade' : faculdade, 'departamento':departamento})


def concluir_registo(request,id):
    ''' Página que é mostrada ao utilizador quando faz um registo na plataforma '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Docente").exists():
            u = "Docente"
        elif user.groups.filter(name = "Administrador").exists():
            u = "Administrador"
        elif user.groups.filter(name = "Funcionário").exists():
            u = "Funcionário"
        else:
            u=""   
    else:
        u=""
    if id == 1:
        participante="True"
    elif id == 0:
        participante="False"
    elif id == 2:
        participante="Admin"   
    return render(request=request,
                  template_name="main/concluir_registo.html",
                  context={'participante': participante, 'u': u})


class consultar_utilizadores(SingleTableMixin, FilterView):
    ''' Consultar todos os utilizadores com as funcionalidades dos filtros '''
    table_class = UtilizadoresTable
    template_name = 'main/consultar_utilizadores.html'
    filterset_class = UtilizadoresFilter
    table_pagination = {
        'per_page': 10
    }

    def dispatch(self, request, *args, **kwargs):
        user_check_var = user_check(
            request=request, user_profile=[Administrador])
        if not user_check_var.get('exists'):
            return user_check_var.get('render')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(SingleTableMixin, self).get_context_data(**kwargs)
        table = self.get_table(**self.get_table_kwargs())
        table.request = self.request
        table.fixed = True
        context[self.get_context_table_name(table)] = table
        return context



def validar_p(request, id,pedidoid): 
    ''' Validar um utilizador na pagina consultar utilizadores '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Funcionário").exists():
            u = "Funcionário"   
        else:
            return redirect('main:mensagem',3) 
    else:
        return redirect('main:mensagem',3) 
        
    try:
        u = Utilizador.objects.get(id = id)
        u.valido = 'True'           
        u.save()   
        subject = 'Validação do pedido na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT'
        message = 'Caro(a) '+u.first_name+"\n\n"
        message+='O seu pedido na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT foi bem sucedido!'+",\n\n"
        message+='Equipa do Conselho Pedagógico da Ualg'
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [u.email,]
        send_mail( subject, message, email_from, recipient_list )
    except SMTPException as e:
     print(f"Failed to send email: {e}")
      

    if 'tablePedidos' not in request.session:
        return redirect('main:tablePedidos')
    else:    
        return HttpResponseRedirect(request.session['tablePedidos'])
    
def validar_utilizador(request, id): 
    ''' Validar um utilizador na pagina consultar utilizadores '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Administrador").exists():
            u = "Administrador"   
        else:
            return redirect('main:mensagem',3) 
    else:
        return redirect('main:mensagem',3) 
        
    try:
        u = Utilizador.objects.get(id = id)
        u.valido = 'True'           
        u.save()   
        subject = 'Validação do registo na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT'
        message = 'Caro(a) '+u.first_name+"\n\n"
        message+='O seu registo na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT foi bem sucedido!'+",\n\n"
        message+='Equipa do Conselho Pedagógico da Ualg'
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [u.email,]
        send_mail( subject, message, email_from, recipient_list )
    except SMTPException as e:
     print(f"Failed to send email: {e}")
      

    if 'consultar_utilizadores' not in request.session:
        return redirect('main:consultar-utilizadores')
    else:    
        return HttpResponseRedirect(request.session['consultar_utilizadores'])



def rejeitar_p(request, id,pedidoid): 
    ''' Funcionalidade de rejeitar um utilizador na pagina de consultar utilizadores '''
    if request.user.is_authenticated:    
        user = get_user(request)
        print(user)
        if user.groups.filter(name = "Funcionário").exists():
            u = "Funcionário"   
            print(u)     
        else:
            return redirect('main:mensagem',3) 
    else:
        return redirect('main:mensagem',3) 
        
    try:
        u = Utilizador.objects.get(id = id)
        pedido = Pedido.objects.get(id=pedidoid)
        u.valido = 'Rejeitado'           
        u.save()   
        subject = 'Validação do pedido na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT'
        message = 'Caro(a) '+u.first_name+",\n\n"
        message+='O seu pedido na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT foi rejeitado!'+"\n\n"
        message+=pedido.Rejeitarpedido+"\n\n"
        message+='Equipa do Conselho Pedagógico da Ualg'
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [u.email,]
        send_mail( subject, message, email_from, recipient_list )
    except:
        pass
    if 'tablePedidos' not in request.session:
        return redirect('main:tablePedidos')
    else:    
        return HttpResponseRedirect(request.session['tablePedidos'])
    
def rejeitar_utilizador(request, id): 
    ''' Funcionalidade de rejeitar um utilizador na pagina de consultar utilizadores '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Administrador").exists():
            u = "Administrador"        
        else:
            return redirect('main:mensagem',3) 
    else:
        return redirect('main:mensagem',3) 
        
    try:
        u = Utilizador.objects.get(id = id)
        u.valido = 'Rejeitado'           
        u.save()   
        subject = 'Validação do registo na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT'
        message = 'Caro(a) '+u.first_name+",\n\n"
        message+='O seu registo na plataforma de Gestão de Pedidos do Conselho Pedagógico da FCT foi rejeitado!'+"\n\n"
        message+='Equipa do Conselho Pedagógico da Ualg'
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [u.email,]
        send_mail( subject, message, email_from, recipient_list )
    except:
        pass
    if 'consultar_utilizadores' not in request.session:
        return redirect('main:consultar-utilizadores')
    else:    
        return HttpResponseRedirect(request.session['consultar_utilizadores'])    


def enviar_email_validar(request,nome,id):
    ''' Envio de email quando o utilizador é validado na pagina consultar utilizadores '''  
    msg="A enviar email a "+nome+" a informar que o seu registo foi validado"
    user_check_var = user_check(request=request, user_profile=[Administrador])
    if user_check_var.get('exists') == False: 
        return user_check_var.get('render')
    request.session['consultar_utilizadores'] = request.META.get('HTTP_REFERER', '/')
    return render(request=request,
                  template_name="main/enviar_email_validar.html",
                  context={"msg": msg, "id":id})



def enviar_email_rejeitar(request,nome,id):  
    ''' Envio de email quando o utilizador é rejeitado na pagina consultar utilizadores '''
    msg="A enviar email a "+nome+" a informar que o seu registo foi rejeitado"
    user_check_var = user_check(request=request, user_profile=[Administrador])
    if user_check_var.get('exists') == False: 
        return user_check_var.get('render')
    request.session['consultar_utilizadores'] = request.META.get('HTTP_REFERER', '/')
    return render(request=request,
                  template_name="main/enviar_email_rejeitar.html",
                  context={"msg": msg, "id":id})


def enviar_email_validarpedido(request, nome, id, pedidoid): 
    msg="A enviar email a "+nome+" a informar que o seu pedido foi validado"

    pedido = Pedido.objects.get(id=pedidoid)
    pedido.status ="Concluido"
    pedido.diaFinal= datetime.date.today()
    pedido.save()
    return render(request=request,
                  template_name="main/enviar_email_validarpedido.html",
                  context={"msg": msg, "id":id,"pedidoid":pedidoid})

def enviar_email_rejeitarpedido(request, nome, id, pedidoid): 
    msg="A enviar email a "+nome+" a informar que o seu pedido foi Rejeitado"
    if request.method == 'POST':
        motivo = request.POST.get('motivo_rejeitar')
    pedido = Pedido.objects.get(id=pedidoid)
    pedido.status ="Concluido"
    pedido.diaFinal= datetime.date.today()
    pedido.Rejeitarpedido = motivo
    pedido.save()
    return render(request=request,
                  template_name="main/enviar_email_rejeitarpedido.html",
                  context={"msg": msg, "id":id,"pedidoid":pedidoid})

def mensagem(request, id, *args, **kwargs):
    ''' Template de mensagens informativas/erro/sucesso '''

    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Administrador").exists():
            u = "Administrador"
        elif user.groups.filter(name = "Docente").exists():
            u = "Docente"
        elif user.groups.filter(name = "Colaborador").exists():
            u = "Funcionario"
        else:
            u=""     
    else:
        u = ""


    if id == 400 or id == 500:
        user = get_user(request)
        m = "Erro no servidor"
        tipo = "error"
    elif id == 1:
        user = get_user(request)
        m = "Bem vindo(a) "+user.first_name
        tipo = "info"

    elif id == 2:
        m = "Até á próxima!"
        tipo = "info"

    elif id == 3:
        m = "Registo feito com sucesso!"
        tipo = "sucess"

    elif id == 4:
        m = "É necessário fazer login primeiro"
        tipo = "error"

    elif id == 5:
        m = "Não permitido"
        tipo = "error"
    elif id == 6:
        m = "Senha alterada com sucesso!"
        tipo = "success"    
    elif id == 7:
        m = "Conta apagada com sucesso"
        tipo = "success"   
    elif id == 8:
        m = "Perfil alterado com sucesso"
        tipo = "success" 
    elif id == 9:
        m = "Perfil criado com sucesso"
        tipo = "success" 
    elif id == 10:
        m = "Não existem notificações"
        tipo = "info"
    elif id == 11:
        m = "Esta tarefa deixou de estar atribuída"
        tipo = "error"
    elif id == 12:
        m = "Ainda não é permitido criar inscrições"
        tipo = "error"
    elif id == 13:
        m = "Erro ao apagar dados do utilizador"
        tipo = "error" 
    elif id == 14:
        m = "Não existem mensagens"
        tipo = "info"  
    elif id == 15:
        m = "Este colaborador tem tarefas iniciadas pelo que apenas deverá ser apagado quando estas estiverem concluidas"
        tipo = "info"  
    elif id == 16:
        m = "Para puder apagar a sua conta deverá concluir primeiro as tarefas que estão iniciadas"
        tipo = "info"                 
    elif id == 17:
        m = "A sua disponibilidade foi alterada com sucesso"
        tipo = "success"
    elif id == 18:
        m = "Antes de poder ver dados e estatísticas é preciso configurar um Dia Aberto."
        tipo = "error"
    else:
        m = "Esta pagina não existe"
        tipo = "error"                                     

    
    continuar = "on" 
    if id == 400 or id == 500:
        continuar = "off" 
    return render(request=request,
        template_name="main/mensagem.html", context={'m': m, 'tipo': tipo ,'u': u, 'continuar': continuar,})

def alterar_utilizador_admin(request,id):
    ''' Funcionalidade de o administrador alterar um utilizador '''
    faculdade = Faculdade.objects.all()
    departamento = Departamento.objects.all()
    if request.user.is_authenticated:    
        utilizador_atual = get_user(request)
        if utilizador_atual.groups.filter(name = "Administrador").exists():
            admin = "Administrador"         
        else:
            return redirect('main:mensagem',3) 
    else:
        return redirect('main:mensagem',3)
    
    user = User.objects.get(id=id)
    if user.groups.filter(name = "Administrador").exists():
        tipo=3
        u = "Administrador"
        utilizador_object = Administrador.objects.get(id=user.id)
        utilizador_form = AdministradorAlterarPerfilForm(instance=utilizador_object)
        perfil="Administrador"
    elif user.groups.filter(name = "Docente").exists():
        tipo=2
        u = "Docente"
        utilizador_object = Docente.objects.get(id=user.id)
        utilizador_form = DocenteAlterarPerfilForm(instance=utilizador_object)
        perfil="Docente"
    elif user.groups.filter(name = "Funcionario").exists():
        tipo=1
        u = "Funcionario" 
        utilizador_object = Funcionario.objects.get(id=user.id)
        utilizador_form = FuncionarioAlterarPerfilForm(instance=utilizador_object)
        perfil= "Funcionario"
    else:
        return redirect('main:mensagem',3)     
    
    msg=False
    if request.method == "POST":
        submitted_data = request.POST.copy()
        if tipo == 1:
            form = FuncionarioAlterarPerfilForm(submitted_data,instance=utilizador_object)
            my_group = Group.objects.get(name='Funcionário') 
        elif tipo == 2:
            form = DocenteAlterarPerfilForm(submitted_data,instance=utilizador_object)
            my_group = Group.objects.get(name='Docente')
        elif tipo == 3:
            form = AdministradorAlterarPerfilForm(submitted_data,instance=utilizador_object)
            my_group = Group.objects.get(name='Administrador')    
        else:
            return redirect('main:mensagem',3)   

        email = request.POST.get('email')

        erros=[]


        if email and User.objects.exclude(email=utilizador_object.email).filter(email=email).exists():
            erros.append('O email já existe')
        elif email==None:
            erros.append('O email é inválido')

        if form.is_valid() and len(erros)==0:
            utilizador_form_object = form.save(commit=False)
            if tipo == 2:
                utilizador_form_object.faculdade = submitted_data['faculdade']
                utilizador_form_object.departamento = submitted_data['departamento']
            utilizador_form_object.save()
            return redirect('main:consultar-utilizadores')   
        else:
            msg=True
            return render(request=request,
                          template_name="main/alterar_utilizador_admin.html",
                          context={"form": form, 'perfil': perfil, 'u': admin,'registo' : tipo,'msg': msg, 'erros':erros,'id':id,'faculdade':faculdade,'departamento':departamento})
    else:

        return render(request=request,
                  template_name="main/alterar_utilizador_admin.html",
                  context={"form": utilizador_form, 'perfil': perfil,'u': admin,'registo' : tipo,'msg': msg,'id':id,'faculdade':faculdade,'departamento':departamento})


def mudar_perfil_escolha_admin(request,id):
    '''  Funcionalidade de o administrador alterar o perfil de um dado utilizador 
     Redireciona para uma pagina onde é possível escolher o perfil que quer alterar '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Administrador").exists():
            u = "Administrador"
        else:
            return redirect('main:mensagem',5)   
    else:
        return redirect('main:mensagem',5) 

    user=User.objects.get(id=id)  
    if user.groups.filter(name = "Administrador").exists():
        x = "Administrador"
    elif user.groups.filter(name = "Docente").exists():
        x = "Docente"
    elif user.groups.filter(name = "Funcionario").exists():
        x = "Funcionario" 
    else:
        return redirect('main:mensagem',5)     

    utilizadores = ["Docente", "Funcionario","Administrador"]
    return render(request=request, template_name='main/mudar_perfil_escolha_admin.html', context={"utilizadores": utilizadores,'u': u,'id':id ,'x':x})

def apagar_utilizador(request, id): 
    ''' Apagar um utilizador na pagina consultar utilizadores '''
    if request.user.is_authenticated:    
        user = get_user(request)
        if user.groups.filter(name = "Administrador").exists():
            u = "Administrador"        
        else:
            return redirect('main:mensagem',5) 
    else:
        return redirect('main:mensagem',5)

    user = User.objects.get(id=id)
    # try:
    if user.groups.filter(name = "Docente").exists():
        u = Docente.objects.get(id=id)
    elif user.groups.filter(name = "Administrador").exists():
        u = Administrador.objects.get(id=id)
    elif user.groups.filter(name = "Funcionario").exists():
        u = Funcionario.objects.get(id=id)
        for tarefa in Pedido.objects.filter(Funcionario=u):
            if tarefa.status=="Em Análise":
                return redirect('main:mensagem',14)
            elif tarefa.status=="Concluida":
                tarefa.delete()
            else:    
                tarefa.atribuido="Não atribuido"
                tarefa.Funcionario=None
                tarefa.save()
    u.delete()             

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

def load_departamentos(request):
    ''' Carregar todos os departamentos para uma determinada faculdade '''
    faculdadeid = request.GET.get('faculdade')
    departamentos = Departamento.objects.filter(unidadeorganicaid=faculdadeid).order_by('nome')
    return render(request, 'main/departamento_dropdown_list_options.html', {'departamentos': departamentos})





def load_cursos(request):
    ''' Carregar todos os cursos para uma determinada faculdade '''
    faculdadeid = request.GET.get('faculdade')
    cursos = Curso.objects.filter(unidadeorganicaid=faculdadeid).order_by('nome')
    return render(request, 'main/curso_dropdown_list_options.html', {'cursos': cursos})



def export(request):
    if request.method == 'POST':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pedido_uc.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'PedidoUC'

        # Set bold font style for column names
        bold_font = Font(bold=True)

        # Write header row
        header_row = ['ID', 'UnidadeCurricular', 'HoraInicio', 'HoraFim', 'Descrição', 'Tarefa', 'ID do Pedido']
        worksheet.append(header_row)

        # Apply bold font style to column names
        for cell in worksheet[1]:
            cell.font = bold_font

        # Retrieve data from the database
        pedido_uc_data = PedidoUC.objects.all()

        # Write data rows
        for pedido_uc in pedido_uc_data:
            worksheet.append([ pedido_uc.id, pedido_uc.uc, pedido_uc.descri, pedido_uc.tipo, pedido_uc.tarefa, pedido_uc.regente, pedido_uc.pedido_id])  # Replace with your actual column names
        workbook.save(response)
        return response

    return render(request, 'main/export.html')


def exportHorarios(request):
    if request.method == 'POST':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pedido_Horario.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'PedidoHorario'

        # Set bold font style for column names
        bold_font = Font(bold=True)

        # Write header row
        header_row = ['ID', 'UnidadeCurricular', 'HoraInicio', 'HoraFim', 'Descrição', 'Tarefa', 'Dia', 'ID do Pedido']
        worksheet.append(header_row)

        # Apply bold font style to column names
        for cell in worksheet[1]:
            cell.font = bold_font

        # Retrieve data from the database
        pedido_uc_data = PedidoHorario.objects.all()

        # Write data rows
        for pedido_uc in pedido_uc_data:
            worksheet.append([ pedido_uc.id, pedido_uc.uc, pedido_uc.hora_inicio, pedido_uc.hora_fim, pedido_uc.descri, pedido_uc.tarefa,pedido_uc.dia, pedido_uc.pedido_id])  # Replace with your actual column names
        workbook.save(response)
        return response

    return render(request, 'main/export.html')


def exportOutros(request):
    if request.method == 'POST':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pedido_Outros.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'PedidoOUT'

        # Set bold font style for column names
        bold_font = Font(bold=True)

        # Write header row
        header_row = ['ID', 'Arquivo', 'Pedido ID']
        worksheet.append(header_row)

        # Apply bold font style to column names
        for cell in worksheet[1]:
            cell.font = bold_font

        # Retrieve data from the database
        pedido_uc_data = PedidosOutros.objects.all()

        # Write data rows
        for pedido_uc in pedido_uc_data:
            worksheet.append([ pedido_uc.id, str(pedido_uc.arquivo), pedido_uc.pedido_id])  # Replace with your actual column names
        workbook.save(response)
        return response

    return render(request, 'main/export.html')

def exportSalas(request):
    if request.method == 'POST':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pedido_salas.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'PedidoSalas'

        # Set bold font style for column names
        bold_font = Font(bold=True)

        # Write header row
        header_row = ['ID', 'Edificio', 'Sala', 'Unidade Curricular', 'Hora de Inicio', 'Hora de Fim', 'Descrição', 'Dia', 'Tarefa', 'ID do pedido']
        worksheet.append(header_row)

        # Apply bold font style to column names
        for cell in worksheet[1]:
            cell.font = bold_font

        # Retrieve data from the database
        pedido_uc_data = PedidoSala.objects.all()

        # Write data rows
        for pedido_uc in pedido_uc_data:
            worksheet.append([ pedido_uc.id, pedido_uc.edi, pedido_uc.sal, pedido_uc.uc, pedido_uc.hora_de_inicio, pedido_uc.hora_de_fim, pedido_uc.descri, pedido_uc.dia, pedido_uc.tarefa, pedido_uc.pedido_id])  # Replace with your actual column names
        workbook.save(response)
        return response

    return render(request, 'main/export.html')